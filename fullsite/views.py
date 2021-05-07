from django.shortcuts import render
from .models import detail,Connect_call,upload_detail
import xlwt
import xlrd
import openpyxl
import datetime
from django.contrib.auth.models import User,auth
from django.contrib.auth import authenticate, login

# Create your views here.

def login(request):
    if request.method =="POST":
        print("ENTER")
        uname=request.POST.get('Username')
        upass=request.POST.get('Password')
        # un="sachin"
        # ps="Mahaprabhu@123"
        # print(uname,upass)
        user = authenticate(username=uname, password=upass)
        print(user)
        if user is not None:
            auth.login(request,user)
            return render(request,'fullsite/home.html')
        else:
            return render(request,'fullsite/login.html',{"message":"Wrong Username or Password..","user":user})
    
    return render(request,'fullsite/login.html')


def user_logout(request):
    auth.logout(request)
    return render(request,'fullsite/home.html')

def home(request):
    return render(request,'fullsite/login.html')

def adminPanel(request):
    if request.method =="POST":
        file=request.FILES['file']
        date = datetime.date.today()
        wb = openpyxl.load_workbook(file)
        print(wb.active.title)
        worksheet = wb["Sheet1"]
        print(worksheet)
        sheets = wb.sheetnames
        for x in range(1,worksheet.max_row):
            phone=worksheet.cell(row=x, column=1).value
            save_detail=detail(date=date,phone=phone)
            save_detail.save()
        ud=upload_detail(date=date,All_call=worksheet.max_row,filetype="Missed Call")
        ud.save()
    ud=upload_detail.objects.all()
    new_lst = ud[::-1]
    return render(request,'fullsite/adminpanel.html',{"upload_detail":new_lst})

def missed(request):
    all_details=detail.objects.all()
    new_lst = all_details[::-1]
    return render(request,'fullsite/missed.html',{"all_detail":new_lst})



def Connected(request):
    if request.method == "POST":
        file=request.FILES['file']
        date = datetime.date.today()
        wb = openpyxl.load_workbook(file)
        print(wb.active.title)
        worksheet = wb["Sheet1"]
        print(worksheet)
        sheets = wb.sheetnames
        for x in range(1,worksheet.max_row):
            phone=worksheet.cell(row=x, column=1).value
            save_detail=Connect_call(date=date,phone=phone)
            save_detail.save()
        ud=upload_detail(date=date,All_call=worksheet.max_row,filetype="Connected Call")
        ud.save()
        ud=upload_detail.objects.all()
        new_lst = ud[::-1]
        return render(request,'fullsite/adminpanel.html',{"upload_detail":new_lst})
    return render(request,'fullsite/login.html')

def update(request):
    if request.method =="POST":
        print("POST REQUEST")
        date = request.POST.get('date')
        phone = request.POST.get('phone')
        status = request.POST.get('status')
        remark = request.POST['remark']
        print(remark)
        Cid=request.POST.get('id')
        save_detail=detail(C_id=Cid,date=date,phone=phone,status=status,remark=remark)
        save_detail.save()
    all_details=detail.objects.all()
    new_lst = all_details[::-1]
    return render(request,'fullsite/missed.html',{"all_detail":new_lst})

def register(request):
    if request.method =="POST":
        uname=request.POST.get('username')
        fname=request.POST.get('fname') 
        lname=request.POST.get('lname')
        pass1=request.POST.get('password')
        pass2=request.POST.get('password2')
        email=request.POST.get('email')
        print(uname,pass1,pass2,email)

        if pass1 == pass2:
            ## check wether user is register or not....
            if User.objects.filter(username=uname).exists():
                return render(request,'fullsite/register.html',{"message":"Username already taken.."})
            elif User.objects.filter(email=email).exists():
                return render(request,'fullsite/register.html',{"message":"Username already exists.."})
            else:
                user=User.objects.create_user(username=uname,password=pass1,email=email,first_name=fname,last_name=lname)
                user.save()
                return render(request,'fullsite/register.html',{"message":"User Created"})
        else:
            return render(request,'fullsite/register.html',{"message":"Password does not match..."})
    return render(request,'fullsite/register.html')